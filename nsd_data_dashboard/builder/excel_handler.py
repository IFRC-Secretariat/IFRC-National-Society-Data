"""
Module to generate and manage an Excel document using Openpyxl, including writing data sheets and adding formatting.
"""
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
import ast
import os
import sys
import shutil
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, borders, numbers
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.formatting.rule import CellIsRule, FormulaRule


class ExcelHandler:
    """
    Class to generate and manage an Excel document using Openpyxl, including writing data sheets and adding formatting.

    Parameters
    ----------
    file_path : string (required)
        Path to the Excel document.
    """
    def __init__(self, file_path):
        # Set attributes
        self.file_path = file_path

        # Create a workbook
        self.book = load_workbook(self.file_path)
        self.writer = pd.ExcelWriter(self.file_path, engine='openpyxl')
        self.writer.book = self.book


    def write_data_sheet(self, data, sheet_name, index=True, startrow=0, header=True, header_height=1, overwrite=False, header_styles=None, body_styles=None, alternate_row_background=None, column_widths=None, index_column_widths=None, level_column_widths=None, sorting=False, hidden=None):
        """
        Write data to a sheet in an Excel document, preserving the other sheets.

        Parameters
        ----------
        data : pandas DataFrame (required)
            Data to insert into the sheet.

        sheet_name : string (required)
            Name of the sheet to insert.

        index : bool (default=True)
            Whether or not to write the index to the sheet.

        startrow : integer (default=0)
            Index of the row at which to start applying formatting. E.g. if startrow=2, 2 blank rows will be left above.

        header : bool (default=True)
            Whether or not to write the dataset column names to the sheet.

        overwrite : bool (default=False)
            If True, an existing sheet with the sheet_name will be deleted before insertion. If False, the sheet will be inserted and the name will be appended with a '1'.

        header_styles : dict (default=None)
            Styles to set to the headers, e.g. {'font': Font(name='Calibri', color='FFFFFF', size=11)}.

        body_styles : dict (default=None)
            Styles to set to the body, i.e. any content that is not the header.

        alternate_row_background : list (default=None)
            Set alternating row background colours for readability.

        column_widths : dict (default=None)
            Dict mapping column names to widths.

        index_column_widths : dict (default=None)
            Dict mapping index names to widths.

        level_column_widths : dict of dicts (default=None)
            Dict in the format {column_level: {column_name: width}}. E.g. {1: {'col1': 10, 'col2': 20}} will set the widths of 'col1' in level 1 and 'col2' in level 1 to 10 and 20 respectively.

        sorting : bool (default=False)
            If True, auto-sorting will be applied to the whole sheet.

        hidden : list (default=None)
            List of indexes of columns to hide (first column is indexed at 0).
        """
        # Check if the sheet exists; if it does, delete it and then write the data
        if sheet_name in self.book.get_sheet_names():
            worksheet = self.book.get_sheet_by_name(sheet_name)
            self.book.remove_sheet(worksheet)
        data.to_excel(excel_writer=self.writer, index=index, sheet_name=sheet_name, header=header, startrow=startrow)

        # Calculate the height of the header and index
        worksheet=self.book.get_sheet_by_name(sheet_name)
        header_height = len(data.columns.names) if not index else len(data.columns.names)+1
        index_size = 0 if not index else len(data.index.names)

        # Set sorting dimensions
        if sorting:
            if index:
                sorting = f'A{startrow+header_height}:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'
            else:
                sorting = f'A{startrow+1}:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'

        # Set the column widths
        column_widths_map = {}
        if column_widths:
            column_widths_map.update({list(data.columns).index(column)+index_size: column_widths[column] for column in column_widths})
        if index and index_column_widths:
            column_widths_map.update({list(data.index.names).index(name): index_column_widths[name] for name in index_column_widths})
        if level_column_widths:
            for level, level_x_column_widths in level_column_widths.items():
                for column in level_x_column_widths:
                    column_widths_map.update({(i+index_size): level_x_column_widths[column] for i, x in enumerate(list(data.columns.get_level_values(level))) if x==column})

        # Apply styles
        worksheet = self.format_sheet(worksheet=worksheet,
                                      header_height=header_height,
                                      header_styles=header_styles,
                                      body_styles=body_styles,
                                      alternate_row_background=alternate_row_background,
                                      column_widths=column_widths_map,
                                      sorting=sorting,
                                      hidden=hidden)


    def format_sheet(self, worksheet, startrow=0, header_styles=None, header_height=1, body_styles=None, alternate_row_background=None, column_widths=None, sorting=False, hidden=None):
        """
        Format an Excel sheet.

        worksheet : openpyxl worksheet object (required)
            Openpyxl Worksheet object to apply the formatting to.

        startrow : integer (default=0)
            Index of the row at which to start applying formatting. E.g. if startrow=2, 2 blank rows will be left above.

        header_styles : dict (default=None)
            Styles to set to the headers, e.g. {'font': Font(name='Calibri', color='FFFFFF', size=11)}.

        header_height : int (default=1)
            The number of rows in the header.

        body_styles : dict (default=None)
            Styles to set to the body, i.e. any content that is not the header.

        alternate_row_background : list (default=None)
            Set alternating row background colours for readability.

        column_widths : dict (default=None)
            Widths of the columns, mapping column indexes (starting at 0) to widths. If None, widths are left as default. E.g. {2: 20, 0: 50, 4:10} will set width of the 3rd column to 20, the first column to 50, and the 5th column to 10.

        sorting : bool or string (default=False)
            Area to apply auto-sorting to. If False, sorting will not be applied.

        hidden : list (default=None)
            List of indexes of columns to hide (first column is indexed at 0).
        """
        # Add font and style formatting to the headers
        if header_styles:
            for row in worksheet[worksheet.dimensions][startrow:startrow+header_height]:
                for cell in row:
                    for style_name in header_styles:
                        setattr(cell, style_name, header_styles[style_name])

        # Add styles to the body: body styles and alternating background if set
        if body_styles or alternate_row_background:
            i=1
            for row in worksheet[worksheet.dimensions][startrow+header_height:]:
                for cell in row:
                    if body_styles:
                        for style_name in body_styles:
                            setattr(cell, style_name, body_styles[style_name])
                    if alternate_row_background:
                        cell.fill = PatternFill(start_color=alternate_row_background[i%2], end_color=alternate_row_background[i%2], fill_type = "solid")
                i += 1

        # Set column width for the data columns
        if column_widths:
            for i in column_widths:
                worksheet.column_dimensions[get_column_letter(i+1)].width = column_widths[i]

        # Add worksheet sorting
        if sorting:
            worksheet.auto_filter.ref = sorting

        # Hide columns
        if hidden:
            for column in hidden:
                worksheet.column_dimensions[get_column_letter(column+1)].hidden = True

        return worksheet


    def add_image(self, image_path, sheets, loc, atts=None):
        """
        Write an image to a sheet at a specified location.

        image_path : string (required)
            Path to the image.

        sheets : list or string (required)
            Worksheet name, or list of worksheet names to write the image to.

        loc : string or list (required)
            Cell name to anchor the image to, e.g. 'A1', this will anchor the image to the top left corner of the cell. Alternatively coordinates can be given in the format [x_coord, y_coord].

        atts : dict (default=None)
            Attributes to set to the img object, e.g. {'width': 658, 'height': 150}.
        """
        # Read in the worksheet name(s)
        if not isinstance(sheets, list): sheets = [sheets]

        # Loop through the sheets and get each sheet
        for sheet_name in sheets:
            worksheet = self.book.get_sheet_by_name(sheet_name)

            # Load the image and apply attributes
            img = Image(image_path)
            if atts:
                for att_name in atts:
                    setattr(img, att_name, atts[att_name])

            # Add more precise positioning if loc is given as coordinates
            if isinstance(loc, list) or isinstance(loc, tuple):
                p2e = pixels_to_EMU
                h, w = img.height, img.width
                position = XDRPoint2D(p2e(loc[0]), p2e(loc[1]))
                size = XDRPositiveSize2D(p2e(w), p2e(h))
                img.anchor = AbsoluteAnchor(pos=position, ext=size)

                # Insert the image
                worksheet.add_image(img)

            else:
                # Insert the image
                worksheet.add_image(img, loc)


    def delete_rows(self, sheets, idx, amount=1):
        """
        Delete blank rows from sheets.

        Parameters
        ----------
        sheets : string or list (required)
            Name of the sheet(s) to delete the rows from.

        idx : int (required)
            Index of the row to delete.

        amount int (default=1)
            Number of rows to delete
        """
        # Loop through worksheets
        if not isinstance(sheets, list): sheets = [sheets]
        for sheet in sheets:

            # Read in the worksheet and delete the rows
            worksheet = self.book.get_sheet_by_name(sheet)
            worksheet.delete_rows(idx=idx, amount=amount)


    def hide_columns(self, sheet_name, columns):
        """
        Format an Excel sheet.

        Parameters
        ----------
        sheet_name : string (required)
            Name of the sheet to hide columns in.

        columns : list (required)
            List of indexes of columns to hide (first column is indexed at 0).
        """
        # Get the worksheet
        worksheet = self.book.get_sheet_by_name(sheet_name)

        # Hide the columns
        if columns:
            for column in columns:
                worksheet.column_dimensions[get_column_letter(column+1)].hidden = True


    def order_sheets(self, order, hidden=None):
        """
        Sort the sheets/ tabs in the workbook.

        order : list (required)
            List of sheet names in the required order.

        hidden : list (default=None)
            List of sheets that should be hidden.
        """
        # Create a dict of sheet names to number
        if hidden: order += hidden
        sheet_positions = {order[i]: i for i in range(0, len(order))}

        # Order the sheets with this mapping
        self.book._sheets.sort(key = lambda worksheet: (len(order)+1) if worksheet.title not in sheet_positions else sheet_positions[worksheet.title])

        # Hide sheets if required
        if hidden:
            for sheet_name in hidden:
                worksheet = self.book.get_sheet_by_name(sheet_name)
                worksheet.sheet_state = 'hidden'


    def lock_worksheets(self, sheets, password=None):
        """
        Lock worksheets in the workbook.

        Parameters
        ----------
        sheets : string or list (required)
            Name of the sheet(s) to lock.

        password : string (default=None)
            Password to use to protect the sheets.
        """
        # Loop through worksheets
        if not isinstance(sheets, list): sheets = [sheets]
        for sheet in sheets:

            # Read in the worksheet and set the protection
            worksheet = self.book.get_sheet_by_name(sheet)
            worksheet.protection.sheet = True
            worksheet.protection.enable()
            if password:
                worksheet.protection.password = password

            # Set the following to make sure that the user can still filter the data
            worksheet.protection.autoFilter = False
            worksheet.protection.sort = False
            worksheet.protection.selectUnlockedCells = True


    def save(self):
        """
        Save the document.
        """
        # Save and close the document
        self.book.save(filename=self.file_path)
