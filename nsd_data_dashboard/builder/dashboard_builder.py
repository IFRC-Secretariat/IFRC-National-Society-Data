"""
Module to generate an Excel dashboard with NSD data.

"""
import pandas as pd
import shutil
import yaml
import ast
import re
from datetime import date
import os
import sys
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border, borders
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from nsd_data_dashboard.builder.excel_handler import ExcelHandler


class NSDDashboardBuilder(ExcelHandler):
    """
    Class to generate an Excel document containing data on IFRC National Societies.

    Parameters
    ----------
    save_folder : string (required)
        The location to save the final Excel document and datasets.

    file_name : string (default='NSD Data Dashboard')
        The name to give the generated Excel dashboard file.
    """
    def __init__(self, save_folder, file_name='NSD Data Dashboard'):
        # Set other variables
        self.save_folder = save_folder
        self.formula_prefix = '_xlfn.'

        # Get the static file directory and excel template path
        __location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
        self.static_folder = os.path.join(__location__, 'static/')
        excel_template = os.path.join(self.static_folder, 'dashboard_template.xlsx')

        # Copy the Excel template into save_folder and initiate the parent class, setting the file path as the path to this copy
        self.excel_file_path = f'{self.save_folder}/{file_name}.xlsx'
        shutil.copyfile(excel_template, self.excel_file_path)
        super().__init__(file_path=self.excel_file_path)


    def generate_dashboard(self, indicator_datasets, categories=None, protect_sheets=False, protect_workbook=False, excel_password=None):
        """
        Generate the Excel document containing information on IFRC National Societies.

        Parameters
        ----------
        indicator_datasets : list of Dataset objects (required)
            List of Dataset objects, containing information on datasets to write to the Excel document.

        categories : dict (default=None)
            Dict specifying data categories. Each category will be written as a sheet in the Excel document.
            Dict keys are category names which are used as sheet names, and the values are a list of indicators belonging to that category. Each indicator should include the indicator name and dataset name.

        protect_sheets : boolean or list of strings (default=False)
            List of sheets to protect. If True, all of the sheets in the Excel document will be protected.

        protect_workbook : boolean (default=False)
            If True, the workbook will be protected and the structure will be locked.

        excel_password : string (default=None)
            If protect_sheets is set to True, this password will be used to protect the sheets.
        """
        # Set attributes
        self.indicator_datasets = indicator_datasets

        # Set column order
        self.info_columns_order = ['Value', 'Year', 'Source', 'Type', 'Link', 'Focal point']
        self.info_column_widths = {'Value': 20, 'Year': 10, 'Source': 15, 'Type': 12, 'Link': 15, 'Focal point': 20}
        self.ns_columns = ['National Society name', 'Country', 'ISO3', 'Region']
        self.ns_column_widths = {'National Society name': 40, 'Country': 25, 'ISO3': 10, 'Region': 30}

        # Process the data into a "log" type dataset and write to the Excel document
        self.df_log = self.process_log_data()
        #self.write_log(data=self.df_log)

        # Process a list of indicators and write to the Excel document
        df_indicators = self.process_indicators_list()
        #self.write_indicators_list(data=df_indicators)

        # Loop through categories and write to the Excel document
        category_colours = (('ec1d25', 'f57b6c'), ('f58225', 'f6a973'), ('f8a71a', 'fcc777'), ('D9CF01', 'BDA203'), ('70bf42', 'add68a'), ('01a55e', '63c296'), ('00aeac', '54c5c3'), ('0367b2', '5c88c5'), ('1f419a', '5566ae'), ('5a3094', '7d6dae'), ('a2238e', 'bd7eb3'), ('db327f', 'f24baa'))
        if categories:
            for i, category in enumerate(categories):
                df_category = self.process_category_data(indicators=category['indicators'])
                if not df_category is None:
                    self.write_category_data(data=df_category,
                                             sheet_name=category['name'],
                                             title=category['name'],
                                             description=category['description'],
                                             header_colours=category_colours[i],
                                             tab_colour=category_colours[i][0])

        # Write the update date to the About sheet and add the header image
        self.write_update_date()
        self.add_image(image_path=os.path.join(self.static_folder, 'header.png'),
                       sheets='About',
                       loc='A1',
                       atts={'height': 195, 'width': 810})

        # Set the sheet order
        sheets = ['About']
        if categories:
            sheets += [category['name'] for category in categories]
        #sheets += ['All data', 'List of indicators']
        self.order_sheets(order=sheets)

        # Protect the sheets
        if protect_sheets is True:
            protect_sheets = self.book.sheetnames
        if protect_sheets:
            self.lock_worksheets(sheets=protect_sheets, password=excel_password)

        # Protect the workbook and structure
        if protect_workbook:
            self.book.security = WorkbookProtection(workbookPassword=excel_password, revisionsPassword=excel_password, lockWindows=True, lockStructure=True, lockRevision=True)

        # Print out a final message
        self.save()
        print('Excel document generated and saved to', self.excel_file_path)


    def process_log_data(self):
        """
        Process all of the datasets into a log format.
        """
        # Convert the datasets into a log format and append them together
        log_datasets = []
        for dataset in self.indicator_datasets:
            df_unstack = dataset.data.stack(level=0)
            missing_columns = [column for column in (self.ns_columns+['Indicator']) if column not in df_unstack.index.names]
            if missing_columns:
                raise ValueError(f'Columns missing from dataset {dataset.name}: ', missing_columns)
            log_datasets.append(df_unstack)
        df_log = pd.concat(log_datasets, axis='rows')

        # Sort the dataset, rename columns
        df_log = df_log.reset_index()\
                       .sort_values(by=['Year', 'Indicator', 'Country'])

        # Order columns
        df_log = df_log[self.ns_columns+['Indicator']+self.info_columns_order]

        return df_log


    def process_indicators_list(self):
        """
        Process a dataset containing a list of all the indicators, including data sources.
        """
        # Get the list of indicators from the log data
        indicator_data = []
        for dataset in self.indicator_datasets:
            dataset_indicators = dataset.data.stack(level=0).reset_index().drop_duplicates(subset=['Indicator'])[['Indicator']]
            for column, value in dataset.meta.items():
                dataset_indicators[column] = value
            indicator_data.append(dataset_indicators)
        df_indicators = pd.concat(indicator_data, axis='rows')

        # Rename the datasets
        df_indicators = df_indicators.sort_values(by=['Type', 'Source', 'Indicator'], ascending=[False, True, True])

        # Order columns
        df_indicators = df_indicators[['Indicator', 'Source', 'Type', 'Link', 'Focal point']]

        return df_indicators


    def process_category_data(self, indicators):
        """
        Get the data for a category by pulling the indicators for that category from the different dataets. Process the data.
        """
        # Get the list of indicators for each dataset
        dataset_indicators = {}
        for indicator in indicators:
            if indicator['dataset'] in dataset_indicators:
                dataset_indicators[indicator['dataset']].append(indicator['name'])
            else:
                dataset_indicators[indicator['dataset']] = [indicator['name']]

        # Loop through the datasets and get the indicators for this category
        category_datasets = []
        for dataset in self.indicator_datasets:
            if dataset.name in dataset_indicators:
                category_datasets.append(dataset.data[dataset_indicators[dataset.name]])
        if category_datasets:
            df_category = pd.concat(category_datasets, axis='columns')
        else:
            return

        # Order both column levels by the indicators list, and order the info columns order
        indicator_names = [indicator['name'] for indicator in indicators]
        info_columns = list(df_category.columns.get_level_values(level=1).unique())
        info_columns = {column: self.info_columns_order.index(column) for column in info_columns}
        info_columns = list(dict(sorted(info_columns.items(), key=lambda column: column[1])).keys())
        df_category = df_category.reindex(columns=indicator_names, level=0)\
                                 .reindex(columns=info_columns, level=1)

        # Sort the data and reset the index, and rename the columns name
        df_category = df_category.sort_values(by=['Country'])\
                                  .reset_index()
        df_category.columns = df_category.columns.rename({'Indicator': None})

        return df_category


    def write_log(self, data, sheet_name="All data"):
        """
        Write all of the data to a "log" sheet in the Excel document.

        Parameters
        ----------
        data : pandas DataFrame (required)
            Dataset of processed data to write to the data sheet.

        sheet_name : string (default='All data')
            Name of the Excel sheet.
        """
        # Set the formatting
        side = Side(border_style='thin', color='FFFFFF')
        header_styles = {'font': Font(name='Calibri', bold=True, color='ffffff', size=11),
                         'alignment': Alignment(horizontal="left", vertical="center"),
                         'fill': PatternFill(start_color="567EBB", end_color="567EBB", fill_type = "solid"),
                         'border': Border(left=side, right=side, top=side, bottom=side)}
        body_styles = {'border': Border(left=side, right=side, top=side, bottom=side),
                       'alignment': Alignment(horizontal="left", vertical="top")}
        column_widths = {**self.ns_column_widths, **{'Indicator': 50}, **self.info_column_widths}

        # Write the data to a sheet in the Excel document
        self.write_data_sheet(data=data,
                              sheet_name=sheet_name,
                              index=False, overwrite=True,
                              header_styles=header_styles,
                              body_styles=body_styles,
                              alternate_row_background=['BBCBE3', 'DCE4F1'],
                              column_widths=column_widths,
                              sorting=True)


    def write_indicators_list(self, data, sheet_name="List of indicators"):
        """
        Write a dataset containing the list of indicators a sheet in the Excel document.

        Parameters
        ----------
        data : pandas DataFrame (required)
            Dataset of processed data to write to the data sheet.

        sheet_name : string (default='Indicators')
            Name of the Excel sheet.
        """
        # Set the formatting
        side = Side(border_style='thin', color='D9D9D9')
        header_styles = {'font': Font(name='Calibri', bold=True, color='ffffff', size=11),
                         'alignment': Alignment(horizontal="left", vertical="center"),
                         'fill': PatternFill(start_color="567EBB", end_color="567EBB", fill_type = "solid"),
                         'border': Border(left=side, right=side, top=side, bottom=side)}
        body_styles = {'border': Border(left=side, right=side, top=side, bottom=side),
                       'alignment': Alignment(horizontal="left", vertical="top"),}
        column_widths = {'Indicator': 50, 'Source': 30, 'Type': 25, 'Link': 40, 'Focal point': 30}

        # Write the data to a sheet in the Excel document
        self.write_data_sheet(data=data,
                              sheet_name=sheet_name,
                              index=False, overwrite=True,
                              header_styles=header_styles,
                              body_styles=body_styles,
                              alternate_row_background=['F2F2F2', 'FFFFFF'],
                              column_widths=column_widths,
                              sorting=True)


    def write_category_data(self, data, sheet_name, title=None, description=None, header_colours=None, tab_colour=None):
        """
        Write a dataset containing data from multiple datasets grouped by a category to the Excel document.

        Parameters
        ----------
        data : pandas DataFrame (required)
            Dataset containing the data to be written.

        sheet_name : string (required)
            Name of the Excel sheet.

        title : string (default=None)
            If provided, a title will be written to the Excel sheet.

        description : string (default=None)
            Description of the dataset to be written to the Excel sheet.

        header_colours : string or list (default=None)
            Colours of the headers. If string, that colour will be applied to all headers. If list, colours will be applied alternately.

        tab_colour : string (default=None)
            Colour of the Excel tab.
        """
        # Set the base formatting
        side = Side(border_style='thin', color='D9D9D9')
        body_styles = {'border': Border(left=side, right=side, top=side, bottom=side),
                       'alignment': Alignment(horizontal="left", vertical="top"),}
        header_styles = {'font': Font(name='Calibri', bold=True, color='ffffff', size=11),
                         'alignment': Alignment(horizontal="center", vertical="center"),
                         'border': Border(left=side, right=side, top=side, bottom=side)}

        # Set the start row based on the title and description, and the data
        data = data.set_index(self.ns_columns)
        startrow = 3 if (title and description) else 2 if (title or description) else 0

        # Write the data to a sheet in the Excel document
        self.write_data_sheet(data=data,
                              sheet_name=sheet_name,
                              index=True,
                              overwrite=True,
                              startrow=startrow,
                              header_styles=header_styles,
                              body_styles=body_styles,
                              alternate_row_background=['F2F2F2', 'FFFFFF'],
                              index_column_widths=self.ns_column_widths,
                              level_column_widths={1: self.info_column_widths},
                              sorting=True,
                              hidden=None)
        worksheet = self.book.get_sheet_by_name(sheet_name)

        # Set altnerating backgrounds to the header based on the level 0 columns
        header_height = len(data.columns.names)+1
        index_size = len(data.index.names)
        if header_colours:
            if isinstance(header_colours, str): header_styles = [header_colours]
            indicators = data.columns.get_level_values(0)
            indicator_colour_map = {indicator: header_colours[i%len(header_colours)] for i, indicator in enumerate(indicators.unique())}

            for i, indicator in enumerate(indicators):
                for row in worksheet.iter_rows(min_row=startrow+1, max_row=startrow+header_height, min_col=i+index_size+1, max_col=i+index_size+1):
                    for cell in row:
                        cell.fill = PatternFill(start_color=indicator_colour_map[indicator], end_color=indicator_colour_map[indicator], fill_type = "solid")

        # Set background styles to the index
        no_border = Side(border_style=None)
        index_styles = {'font': Font(name='Calibri', bold=True, color='000000', size=11),
                        'alignment': Alignment(horizontal="center", vertical="center"),
                        'fill': PatternFill(start_color='E7E4CF', end_color='E7E4CF', fill_type = "solid"),
                        'border': Border(left=no_border, right=no_border, top=no_border, bottom=no_border)}
        for row in worksheet.iter_rows(min_row=startrow+header_height, max_row=startrow+header_height, min_col=1, max_col=index_size):
            for cell in row:
                cell.fill = PatternFill(start_color='E7E4CF', end_color='E7E4CF', fill_type = "solid")
                cell.font = Font(name='Calibri', bold=True, color='000000', size=11)

        # Add the title and description
        if title:
            self.write_sheet_title(worksheet=worksheet, title=title)
            if description:
                self.write_sheet_description(worksheet=worksheet, description=description, pos='A2')
        elif description:
            self.write_sheet_description(worksheet=worksheet, description=description, pos='A2')

        # Add the worksheet tab colour
        if tab_colour:
            worksheet.sheet_properties.tabColor = tab_colour


    def write_sheet_title(self, worksheet, title, pos='A1'):
        """
        Write a title to the sheet.
        """
        # Write the text and add styles
        worksheet[pos] = title
        title_styles = {'font': Font(name='Calibri', bold=True, color='000000', size=18)}
        for style_name in title_styles:
            setattr(worksheet[pos], style_name, title_styles[style_name])


    def write_sheet_description(self, worksheet, description, pos='A2', merge_cells=None):
        """
        Write a description to the sheet.
        """
        # Merge cells if required
        if merge_cells:
            row_number = re.findall('\d+$', pos)[0]
            worksheet.merge_cells(f'{pos}:{get_column_letter(merge_cells)}{row_number}')

        # Write the text and add styles
        worksheet[pos] = description
        description_styles = {'font': Font(name='Calibri', bold=False, color='000000', size=14)}
        for style_name in description_styles:
            setattr(worksheet[pos], style_name, description_styles[style_name])


    def write_update_date(self):
        """
        Write a sheet with the publication date to the Excel document.
        """
        # Read in the 'About' worksheet
        worksheet = self.book.get_sheet_by_name('About')

        # Write the date to the 'About' sheet in the Excel doc
        today = date.today().strftime('%d %B %Y')
        worksheet['A3'] = f'Updated: {today}'
